"""
Excel文档解析器模块
Excel document parser module for extracting and processing table data.
"""

import pandas as pd
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
import logging
import openpyxl
import numpy as np
from fiscal_mind.semantic_resolver import SemanticResolver

logger = logging.getLogger(__name__)


# Configuration constants for table detection
HEADER_TEXT_THRESHOLD = 0.5  # Minimum proportion of text cells in a header row
MAX_DESCRIPTION_SEARCH_ROWS = 3  # Maximum rows to search above table for description
MAX_CONSECUTIVE_EMPTY_COLUMNS = 1  # Maximum consecutive empty columns in a table
MAX_HEADER_ROWS = 5  # Maximum number of rows to consider for multi-row headers
MIN_TABLE_ROWS = 3  # Minimum data rows for a valid table (excludes small/temporary data)
MIN_TABLE_COLS = 2  # Minimum columns for a valid table


class TableInfo:
    """表格信息类，存储检测到的表格元数据"""
    
    def __init__(self, data: pd.DataFrame, start_row: int, start_col: int, 
                 description: Optional[str] = None):
        """
        初始化表格信息
        
        Args:
            data: 表格数据DataFrame
            start_row: 表格起始行号（0-based）
            start_col: 表格起始列号（0-based）
            description: 表格描述文本
        """
        self.data = data
        self.start_row = start_row
        self.start_col = start_col
        self.description = description
        self.end_row = start_row + len(data)
        self.end_col = start_col + len(data.columns)
    
    def __repr__(self):
        return (f"TableInfo(shape={self.data.shape}, "
                f"position=({self.start_row},{self.start_col}), "
                f"description='{self.description}')")


class TableDetector:
    """表格检测器，用于在sheet中检测多个表格"""
    
    @staticmethod
    def _collect_potential_headers_with_merges(ws, data_array: List[List[Any]], row_idx: int, 
                                                start_col: int, max_col: int,
                                                merged_cache: Optional[Dict[str, Any]] = None) -> Tuple[List[Any], int]:
        """
        收集潜在的表头，考虑合并单元格
        
        Args:
            ws: openpyxl工作表对象
            data_array: 数据数组
            row_idx: 起始行索引
            start_col: 起始列索引
            max_col: 最大列数
            merged_cache: 合并单元格缓存（可选）
            
        Returns:
            (表头列表, 结束列索引)
        """
        potential_headers = []
        temp_col = start_col
        empty_count = 0
        consecutive_empty = 0
        
        while temp_col < max_col:
            # 使用合并单元格值
            value = TableDetector._get_merged_cell_value(ws, row_idx, temp_col, merged_cache)
            
            if pd.notna(value) and str(value).strip():
                # 遇到非空单元格
                # 如果之前有空列但不超过阈值，填充None
                if empty_count > 0:
                    for _ in range(empty_count):
                        potential_headers.append(None)
                potential_headers.append(value)
                empty_count = 0
                consecutive_empty = 0
                temp_col += 1
            else:
                # 遇到空单元格
                empty_count += 1
                consecutive_empty += 1
                temp_col += 1
                # 如果有连续空列，认为表格结束
                if consecutive_empty > MAX_CONSECUTIVE_EMPTY_COLUMNS:
                    break
        
        # 如果没有找到任何表头，返回空列表和start_col
        if not potential_headers:
            return [], start_col
        
        end_col = start_col + len(potential_headers) - 1
        return potential_headers, end_col
    
    @staticmethod
    def _build_merged_cell_cache(ws) -> Dict[str, Any]:
        """
        构建合并单元格的缓存字典，提高查询性能
        
        Args:
            ws: openpyxl工作表对象
            
        Returns:
            字典，键为单元格坐标，值为合并区域左上角单元格的值
        """
        merged_cache = {}
        for merged_range in ws.merged_cells.ranges:
            # 获取左上角单元格的值
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left_value = top_left_cell.value
            
            # 为合并区域内的所有单元格建立缓存
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_coord = ws.cell(row=row, column=col).coordinate
                    merged_cache[cell_coord] = top_left_value
        
        return merged_cache
    
    @staticmethod
    def _get_merged_cell_value(ws, row_idx: int, col_idx: int, 
                               merged_cache: Optional[Dict[str, Any]] = None) -> Any:
        """
        获取合并单元格的值（从合并区域的左上角单元格获取）
        
        Args:
            ws: openpyxl工作表对象
            row_idx: 行索引（0-based）
            col_idx: 列索引（0-based）
            merged_cache: 合并单元格缓存（可选，用于性能优化）
            
        Returns:
            单元格的值
        """
        cell = ws.cell(row=row_idx + 1, column=col_idx + 1)  # openpyxl使用1-based索引
        
        # 如果有缓存，直接从缓存获取
        if merged_cache is not None and cell.coordinate in merged_cache:
            return merged_cache[cell.coordinate]
        
        # 如果没有缓存，检查单元格是否是合并单元格的一部分
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 返回合并区域左上角单元格的值
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        
        return cell.value
    
    @staticmethod
    def _detect_multi_row_headers(ws, data_array: List[List[Any]], start_row: int, 
                                   start_col: int, end_col: int,
                                   merged_cache: Optional[Dict[str, Any]] = None) -> Tuple[List[str], int]:
        """
        检测和处理多行表头，包括合并单元格
        
        Args:
            ws: openpyxl工作表对象
            data_array: 数据数组
            start_row: 起始行索引
            start_col: 起始列索引
            end_col: 结束列索引
            merged_cache: 合并单元格缓存（可选）
            
        Returns:
            (合并后的表头列表, 表头行数)
        """
        max_row = len(data_array)
        header_rows = []
        num_cols = end_col - start_col + 1
        
        # 检查后续几行是否也是表头的一部分
        current_row = start_row
        header_row_count = 0
        
        # 收集潜在的表头行（使用合并单元格值）
        for offset in range(MAX_HEADER_ROWS):
            if current_row + offset >= max_row:
                break
            
            # 读取这一行的值，考虑合并单元格
            row_values = []
            for col_offset in range(num_cols):
                col_idx = start_col + col_offset
                value = TableDetector._get_merged_cell_value(ws, current_row + offset, col_idx, merged_cache)
                row_values.append(value)
            
            # 如果这一行大部分是文本且不是数据行，可能是表头的一部分
            if TableDetector._is_likely_header_row(row_values):
                header_rows.append(row_values)
                header_row_count += 1
            else:
                # 如果遇到数据行，停止
                break
        
        # 如果没有检测到表头行，返回空列表和0
        if header_row_count == 0:
            return [], 0
        
        # 如果只有一行表头，直接返回
        if header_row_count == 1:
            return header_rows[0], 1
        
        # 合并多行表头
        merged_headers = []
        for col_offset in range(num_cols):
            header_parts = []
            previous_value = None
            
            for row_offset in range(header_row_count):
                value = header_rows[row_offset][col_offset]
                
                if pd.notna(value) and str(value).strip():
                    value_str = str(value).strip()
                    # 避免重复添加相同的值（合并单元格跨行时会重复）
                    if value_str != previous_value:
                        header_parts.append(value_str)
                        previous_value = value_str
            
            # 使用连字符连接多层表头
            if header_parts:
                merged_header = '-'.join(header_parts)
                merged_headers.append(merged_header)
            else:
                # 如果该列在所有表头行中都是空的，使用None
                merged_headers.append(None)
        
        return merged_headers, header_row_count
    
    @staticmethod
    def _is_likely_header_row(row_data: List[Any]) -> bool:
        """
        判断是否可能是表头行
        
        Args:
            row_data: 行数据列表
            
        Returns:
            是否可能是表头
        """
        # 过滤掉空值
        non_empty = [v for v in row_data if pd.notna(v) and str(v).strip()]
        
        # 空行不是表头
        if len(non_empty) == 0:
            return False
        
        # 至少有2个非空值才可能是表头
        if len(non_empty) < 2:
            return False
        
        # 检查是否大部分是字符串且不是数字
        text_count = 0
        for v in non_empty:
            try:
                # 如果能转换为数字，可能不是表头
                float(str(v).replace('%', '').replace(',', ''))
            except (ValueError, AttributeError):
                text_count += 1
        
        # 至少50%是文本才可能是表头
        return text_count >= len(non_empty) * HEADER_TEXT_THRESHOLD
    
    @staticmethod
    def _is_data_row(row_data: List[Any], header_count: int) -> bool:
        """
        判断是否是数据行
        
        Args:
            row_data: 行数据列表
            header_count: 表头列数
            
        Returns:
            是否是数据行
        """
        non_empty = [v for v in row_data[:header_count] if pd.notna(v) and str(v).strip()]
        # 至少有一个非空值
        return len(non_empty) > 0
    
    @staticmethod
    def detect_tables(workbook_path: str, sheet_name: str) -> List[TableInfo]:
        """
        检测sheet中的所有表格
        
        Args:
            workbook_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            检测到的表格信息列表
        """
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        ws = wb[sheet_name]
        
        # 读取所有单元格数据到二维数组
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 读取数据到numpy数组
        data_array = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_data = [cell.value for cell in row]
            data_array.append(row_data)
        
        # 构建合并单元格缓存以提高性能
        merged_cache = TableDetector._build_merged_cell_cache(ws)
        
        tables = []
        processed_cells = set()  # 存储已处理的单元格坐标 (row, col)
        
        # 遍历每一行查找表头
        for row_idx in range(max_row):
            row_data = data_array[row_idx]
            
            # 在这一行中查找可能的表头区域
            col_idx = 0
            while col_idx < max_col:
                # 跳过已处理的单元格
                if (row_idx, col_idx) in processed_cells:
                    col_idx += 1
                    continue
                
                # 跳过空单元格（使用合并单元格值检查）
                cell_value = TableDetector._get_merged_cell_value(ws, row_idx, col_idx, merged_cache)
                if pd.isna(cell_value) or not str(cell_value).strip():
                    col_idx += 1
                    continue
                
                # 找到非空单元格，检查是否是表头的开始
                start_col = col_idx
                
                # 收集连续的非空单元格作为潜在表头（考虑合并单元格）
                potential_headers, end_col = TableDetector._collect_potential_headers_with_merges(
                    ws, data_array, row_idx, start_col, max_col, merged_cache
                )
                
                # 检查这些单元格是否像表头
                if len(potential_headers) >= MIN_TABLE_COLS and TableDetector._is_likely_header_row(potential_headers):
                    # 使用多行表头检测
                    headers, header_row_count = TableDetector._detect_multi_row_headers(
                        ws, data_array, row_idx, start_col, end_col, merged_cache
                    )
                    header_count = len(headers)
                    
                    # 查找数据行
                    data_rows = []
                    description = None
                    
                    # 检查表头上方是否有描述（向上最多查找3行）
                    for desc_idx in range(max(0, row_idx - MAX_DESCRIPTION_SEARCH_ROWS), row_idx):
                        desc_row = data_array[desc_idx]
                        # 只检查相同列区域的描述（额外检查一列以防描述略微超出表格）
                        for check_col in range(start_col, min(end_col + 1, max_col)):
                            if pd.notna(desc_row[check_col]) and str(desc_row[check_col]).strip():
                                cell_str = str(desc_row[check_col]).strip()
                                # 如果文本较长且包含中文或特殊字符，可能是描述
                                if len(cell_str) > 3 and ('表' in cell_str or '数据' in cell_str or '：' in cell_str or ':' in cell_str):
                                    description = cell_str
                                    # 标记描述所在的行为已处理
                                    for mark_col in range(max_col):
                                        processed_cells.add((desc_idx, mark_col))
                                    break
                        if description:
                            break
                    
                    # 提取数据行（从表头之后开始）
                    data_start_row = row_idx + header_row_count
                    for data_row_idx in range(data_start_row, max_row):
                        data_row = data_array[data_row_idx]
                        data_slice = data_row[start_col:start_col + header_count]
                        
                        if TableDetector._is_data_row(data_slice, header_count):
                            data_rows.append(data_slice)
                            # 标记这些单元格为已处理
                            for mark_col in range(start_col, start_col + header_count):
                                processed_cells.add((data_row_idx, mark_col))
                        else:
                            # 遇到空行，表格结束
                            if len(data_rows) > 0:
                                break
                    
                    # 过滤小表格：只有数据行数 >= MIN_TABLE_ROWS 且列数 >= MIN_TABLE_COLS 的表格才被认为是有效表格
                    if len(data_rows) >= MIN_TABLE_ROWS and header_count >= MIN_TABLE_COLS:
                        # 创建DataFrame
                        df = pd.DataFrame(data_rows, columns=headers)
                        
                        # 创建TableInfo
                        table_info = TableInfo(
                            data=df,
                            start_row=row_idx,
                            start_col=start_col,
                            description=description
                        )
                        tables.append(table_info)
                        
                        logger.info(f"检测到表格: 位置({row_idx},{start_col}), "
                                  f"大小({len(data_rows)}x{header_count}), "
                                  f"表头行数={header_row_count}")
                    else:
                        logger.debug(f"跳过小表格: 位置({row_idx},{start_col}), "
                                   f"大小({len(data_rows)}x{header_count}) - "
                                   f"不满足最小尺寸要求(行>={MIN_TABLE_ROWS}, 列>={MIN_TABLE_COLS})")
                    
                    # 标记表头行的这些单元格为已处理
                    for header_row_offset in range(header_row_count):
                        for mark_col in range(start_col, start_col + header_count):
                            processed_cells.add((row_idx + header_row_offset, mark_col))
                    
                    # 移动到这个区域之后
                    col_idx = end_col + 1
                else:
                    col_idx += 1
        
        wb.close()
        return tables


class ExcelDocument:
    """表示单个Excel文档及其内容"""
    
    def __init__(self, file_path: str, detect_multiple_tables: bool = False, 
                 semantic_resolver: Optional[SemanticResolver] = None):
        """
        初始化Excel文档
        
        Args:
            file_path: Excel文件路径
            detect_multiple_tables: 是否检测多表格，默认False保持向后兼容
            semantic_resolver: 语义解析器实例（可选）
        """
        self.file_path = Path(file_path)
        self.file_name = self.file_path.name
        self.sheets: Dict[str, pd.DataFrame] = {}
        self.detect_multiple_tables = detect_multiple_tables
        self.multi_tables: Dict[str, List[TableInfo]] = {}  # 存储多表格信息
        self.semantic_resolver = semantic_resolver or SemanticResolver()
        self._load_document()
    
    def _load_document(self):
        """加载Excel文档的所有工作表"""
        try:
            # 读取所有工作表
            excel_file = pd.ExcelFile(self.file_path)
            
            for sheet_name in excel_file.sheet_names:
                if self.detect_multiple_tables:
                    # 使用多表格检测
                    tables = TableDetector.detect_tables(str(self.file_path), sheet_name)
                    
                    if len(tables) == 0:
                        # 如果没有检测到表格，使用默认方式读取
                        self.sheets[sheet_name] = pd.read_excel(
                            excel_file, 
                            sheet_name=sheet_name
                        )
                        logger.info(f"工作表 '{sheet_name}' 未检测到表格，使用默认方式读取")
                    elif len(tables) == 1:
                        # 只有一个表格，直接使用
                        self.sheets[sheet_name] = tables[0].data
                        self.multi_tables[sheet_name] = tables
                        logger.info(f"工作表 '{sheet_name}' 检测到1个表格 (位置: 行{tables[0].start_row}, 列{tables[0].start_col})")
                    else:
                        # 多个表格，使用第一个作为主表格（向后兼容）
                        self.sheets[sheet_name] = tables[0].data
                        self.multi_tables[sheet_name] = tables
                        logger.info(f"工作表 '{sheet_name}' 检测到{len(tables)}个表格")
                else:
                    # 默认方式：读取整个sheet
                    self.sheets[sheet_name] = pd.read_excel(
                        excel_file, 
                        sheet_name=sheet_name
                    )
            
            logger.info(f"成功加载文档: {self.file_name}, 包含 {len(self.sheets)} 个工作表")
        except Exception as e:
            logger.error(f"加载文档失败 {self.file_path}: {str(e)}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """获取所有工作表名称"""
        return list(self.sheets.keys())
    
    def get_sheet(self, sheet_name: str, use_semantic: bool = True) -> Optional[pd.DataFrame]:
        """
        获取指定工作表的数据
        
        Args:
            sheet_name: 工作表名称
            use_semantic: 是否使用语义匹配（如果精确匹配失败）
            
        Returns:
            DataFrame，如果未找到则返回None
        """
        # First try exact match
        if sheet_name in self.sheets:
            return self.sheets.get(sheet_name)
        
        # If exact match fails and semantic is enabled, try semantic matching
        if use_semantic and self.semantic_resolver:
            matched_sheet = self.semantic_resolver.find_sheet_by_semantic(
                list(self.sheets.keys()), 
                sheet_name
            )
            if matched_sheet:
                logger.info(f"Semantic match: '{sheet_name}' -> '{matched_sheet}'")
                return self.sheets.get(matched_sheet)
        
        return None
    
    def get_sheet_tables(self, sheet_name: str) -> Optional[List[TableInfo]]:
        """
        获取工作表中检测到的所有表格
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            表格信息列表，如果未启用多表格检测或不存在则返回None
        """
        return self.multi_tables.get(sheet_name)
    
    def get_table_by_index(self, sheet_name: str, table_index: int) -> Optional[pd.DataFrame]:
        """
        获取工作表中指定索引的表格数据
        
        Args:
            sheet_name: 工作表名称
            table_index: 表格索引（0-based）
            
        Returns:
            表格DataFrame，如果不存在则返回None
        """
        tables = self.multi_tables.get(sheet_name)
        if tables and 0 <= table_index < len(tables):
            return tables[table_index].data
        return None
    
    def get_table_info(self, sheet_name: str, table_index: int) -> Optional[TableInfo]:
        """
        获取工作表中指定索引的表格完整信息
        
        Args:
            sheet_name: 工作表名称
            table_index: 表格索引（0-based）
            
        Returns:
            TableInfo对象，如果不存在则返回None
        """
        tables = self.multi_tables.get(sheet_name)
        if tables and 0 <= table_index < len(tables):
            return tables[table_index]
        return None
    
    def get_sheet_summary(self, sheet_name: str) -> Dict[str, Any]:
        """
        获取工作表摘要信息
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            包含工作表统计信息的字典
        """
        df = self.get_sheet(sheet_name)
        if df is None:
            return {}
        
        summary = {
            "sheet_name": sheet_name,
            "rows": len(df),
            "columns": len(df.columns),
            "column_names": df.columns.tolist(),
            "column_types": df.dtypes.astype(str).to_dict(),
            "has_null": df.isnull().any().to_dict(),
            "memory_usage": f"{df.memory_usage(deep=True).sum() / 1024:.2f} KB"
        }
        
        # 如果启用了多表格检测，添加表格信息
        if self.detect_multiple_tables and sheet_name in self.multi_tables:
            tables = self.multi_tables[sheet_name]
            summary["num_tables"] = len(tables)
            summary["tables"] = []
            for i, table_info in enumerate(tables):
                table_summary = {
                    "index": i,
                    "position": f"行{table_info.start_row},列{table_info.start_col}",
                    "shape": table_info.data.shape,
                    "description": table_info.description,
                    "columns": table_info.data.columns.tolist()
                }
                summary["tables"].append(table_summary)
        
        return summary
    
    def get_document_summary(self) -> Dict[str, Any]:
        """
        获取整个文档的摘要信息
        
        Returns:
            包含文档统计信息的字典
        """
        total_rows = sum(len(df) for df in self.sheets.values())
        total_cols = sum(len(df.columns) for df in self.sheets.values())
        
        summary = {
            "file_name": self.file_name,
            "file_path": str(self.file_path),
            "num_sheets": len(self.sheets),
            "sheet_names": self.get_sheet_names(),
            "total_rows": total_rows,
            "total_columns": total_cols,
            "sheets_summary": {
                name: self.get_sheet_summary(name) 
                for name in self.get_sheet_names()
            }
        }
        
        return summary
    
    def search_value(self, value: Any, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        在工作表中搜索特定值
        
        Args:
            value: 要搜索的值
            sheet_name: 指定工作表名称，如果为None则搜索所有工作表
            
        Returns:
            包含搜索结果的列表
        """
        results = []
        sheets_to_search = {sheet_name: self.sheets[sheet_name]} if sheet_name else self.sheets
        
        for name, df in sheets_to_search.items():
            # 搜索包含该值的单元格
            for col in df.columns:
                mask = df[col].astype(str).str.contains(str(value), case=False, na=False)
                if mask.any():
                    matching_rows = df[mask]
                    for idx, row in matching_rows.iterrows():
                        results.append({
                            "sheet": name,
                            "row": idx,
                            "column": col,
                            "value": row[col],
                            "row_data": row.to_dict()
                        })
        
        return results
    
    def get_column_data(self, column_name: str, sheet_name: str) -> Optional[pd.Series]:
        """
        获取指定列的数据
        
        Args:
            column_name: 列名
            sheet_name: 工作表名称
            
        Returns:
            列数据Series，如果不存在则返回None
        """
        df = self.get_sheet(sheet_name)
        if df is not None and column_name in df.columns:
            return df[column_name]
        return None
    
    def filter_rows(self, sheet_name: str, conditions: Dict[str, Any]) -> Optional[pd.DataFrame]:
        """
        根据条件过滤行（保留向后兼容，等于操作）
        
        Args:
            sheet_name: 工作表名称
            conditions: 过滤条件字典，键为列名，值为过滤值
            
        Returns:
            过滤后的DataFrame
        """
        df = self.get_sheet(sheet_name)
        if df is None:
            return None
        
        filtered_df = df.copy()
        for col, value in conditions.items():
            if col in df.columns:
                filtered_df = filtered_df[filtered_df[col] == value]
        
        return filtered_df
    
    def filter_rows_advanced(self, sheet_name: str, filters: List[Dict[str, Any]], 
                            use_semantic: bool = True) -> Optional[pd.DataFrame]:
        """
        高级行过滤，支持多种比较操作符和语义列名匹配
        
        Args:
            sheet_name: 工作表名称
            filters: 过滤条件列表，每个条件是一个字典:
                    {
                        'column': '列名',
                        'operator': '操作符',  # ==, !=, >, <, >=, <=, between, in, contains
                        'value': 值 或 [min, max] (for between) 或 [值列表] (for in)
                    }
            use_semantic: 是否使用语义列名匹配（当精确匹配失败时）
        
        Returns:
            过滤后的DataFrame
            
        Example:
            filters = [
                {'column': '销售额', 'operator': '>', 'value': 1000000},
                {'column': '区域', 'operator': 'in', 'value': ['华东', '华南']},
                {'column': '日期', 'operator': 'between', 'value': ['2023-01-01', '2023-12-31']}
            ]
        """
        df = self.get_sheet(sheet_name)
        if df is None:
            return None
        
        filtered_df = df.copy()
        
        for filter_cond in filters:
            col = filter_cond.get('column')
            operator = filter_cond.get('operator', '==')
            value = filter_cond.get('value')
            
            # Try to find the column (exact or semantic match)
            actual_col = col
            if col not in df.columns:
                if use_semantic and self.semantic_resolver:
                    # Try semantic matching
                    matched_cols = self.semantic_resolver.find_column_by_semantic(df, col)
                    if matched_cols:
                        actual_col = matched_cols[0]
                        logger.info(f"Semantic column match: '{col}' -> '{actual_col}'")
                    else:
                        logger.warning(f"列 '{col}' 不存在且未找到语义匹配，跳过此过滤条件")
                        continue
                else:
                    logger.warning(f"列 '{col}' 不存在，跳过此过滤条件")
                    continue
            
            try:
                if operator == '==':
                    # Support fuzzy value matching for categorical data
                    if use_semantic and pd.api.types.is_object_dtype(filtered_df[actual_col]):
                        # Normalize values for comparison
                        normalized_value = self.semantic_resolver.normalize_value(str(value))
                        filtered_df = filtered_df[
                            filtered_df[actual_col].apply(
                                lambda x: self.semantic_resolver.normalize_value(str(x)) == normalized_value
                            )
                        ]
                    else:
                        filtered_df = filtered_df[filtered_df[actual_col] == value]
                elif operator == '!=':
                    filtered_df = filtered_df[filtered_df[actual_col] != value]
                elif operator == '>':
                    filtered_df = filtered_df[filtered_df[actual_col] > value]
                elif operator == '<':
                    filtered_df = filtered_df[filtered_df[actual_col] < value]
                elif operator == '>=':
                    filtered_df = filtered_df[filtered_df[actual_col] >= value]
                elif operator == '<=':
                    filtered_df = filtered_df[filtered_df[actual_col] <= value]
                elif operator == 'between':
                    if isinstance(value, (list, tuple)) and len(value) == 2:
                        filtered_df = filtered_df[
                            (filtered_df[actual_col] >= value[0]) & (filtered_df[actual_col] <= value[1])
                        ]
                    else:
                        logger.warning(f"between操作符需要[min, max]格式的值")
                elif operator == 'in':
                    if isinstance(value, (list, tuple)):
                        # Support fuzzy value matching for categorical data
                        if use_semantic and pd.api.types.is_object_dtype(filtered_df[actual_col]):
                            normalized_values = [self.semantic_resolver.normalize_value(str(v)) for v in value]
                            filtered_df = filtered_df[
                                filtered_df[actual_col].apply(
                                    lambda x: self.semantic_resolver.normalize_value(str(x)) in normalized_values
                                )
                            ]
                        else:
                            filtered_df = filtered_df[filtered_df[actual_col].isin(value)]
                    else:
                        logger.warning(f"in操作符需要列表格式的值")
                elif operator == 'contains':
                    filtered_df = filtered_df[
                        filtered_df[actual_col].astype(str).str.contains(str(value), case=False, na=False)
                    ]
                else:
                    logger.warning(f"不支持的操作符: {operator}")
            except Exception as e:
                logger.error(f"过滤条件应用失败 {filter_cond}: {str(e)}")
                continue
        
        return filtered_df
    
    def sort_rows(self, sheet_name: str, sort_by: List[Dict[str, Any]]) -> Optional[pd.DataFrame]:
        """
        对工作表行进行排序
        
        Args:
            sheet_name: 工作表名称
            sort_by: 排序条件列表，每个条件是一个字典:
                    {
                        'column': '列名',
                        'ascending': True/False  # True为升序，False为降序
                    }
        
        Returns:
            排序后的DataFrame
            
        Example:
            sort_by = [
                {'column': '销售额', 'ascending': False},  # 按销售额降序
                {'column': '日期', 'ascending': True}       # 再按日期升序
            ]
        """
        df = self.get_sheet(sheet_name)
        if df is None:
            return None
        
        # 提取列名和排序方向
        columns = []
        ascending = []
        
        for sort_cond in sort_by:
            col = sort_cond.get('column')
            asc = sort_cond.get('ascending', True)
            
            if col in df.columns:
                columns.append(col)
                ascending.append(asc)
            else:
                logger.warning(f"列 '{col}' 不存在，跳过此排序条件")
        
        if not columns:
            return df.copy()
        
        try:
            sorted_df = df.sort_values(by=columns, ascending=ascending)
            return sorted_df
        except Exception as e:
            logger.error(f"排序失败: {str(e)}")
            return df.copy()


class TableJoiner:
    """表格关联工具类，支持多表join操作"""
    
    @staticmethod
    def join_tables(left_df: pd.DataFrame, right_df: pd.DataFrame, 
                   left_on: str, right_on: str, 
                   how: str = 'inner') -> Optional[pd.DataFrame]:
        """
        关联两个表格
        
        Args:
            left_df: 左表DataFrame
            right_df: 右表DataFrame
            left_on: 左表关联键
            right_on: 右表关联键
            how: 关联方式 ('inner', 'left', 'right', 'outer')
            
        Returns:
            关联后的DataFrame
        """
        try:
            result = pd.merge(
                left_df, 
                right_df, 
                left_on=left_on, 
                right_on=right_on, 
                how=how,
                suffixes=('_left', '_right')
            )
            return result
        except Exception as e:
            logger.error(f"表格关联失败: {str(e)}")
            return None
    
    @staticmethod
    def vlookup(lookup_df: pd.DataFrame, table_df: pd.DataFrame, 
               lookup_column: str, return_column: str,
               match_column: str) -> pd.Series:
        """
        模拟Excel的VLOOKUP功能
        
        Args:
            lookup_df: 需要查找的DataFrame
            table_df: 查找表DataFrame
            lookup_column: lookup_df中用于匹配的列
            match_column: table_df中用于匹配的列
            return_column: table_df中要返回的列
            
        Returns:
            查找结果Series
        """
        try:
            # 创建查找字典
            lookup_dict = table_df.set_index(match_column)[return_column].to_dict()
            # 应用查找
            result = lookup_df[lookup_column].map(lookup_dict)
            return result
        except Exception as e:
            logger.error(f"VLOOKUP失败: {str(e)}")
            return pd.Series()


class ExcelParser:
    """Excel文档解析器，支持处理多个文档"""
    
    def __init__(self, detect_multiple_tables: bool = True, llm=None):
        """
        初始化解析器
        
        Args:
            detect_multiple_tables: 是否检测sheet中的多个表格
            llm: 语言模型实例（可选，用于语义匹配）
        """
        self.documents: Dict[str, ExcelDocument] = {}
        self.joiner = TableJoiner()
        self.detect_multiple_tables = detect_multiple_tables
        self.semantic_resolver = SemanticResolver(llm=llm)
        self.llm = llm
    
    def load_document(self, file_path: str, detect_multiple_tables: Optional[bool] = None) -> ExcelDocument:
        """
        加载Excel文档
        
        Args:
            file_path: Excel文件路径
            detect_multiple_tables: 是否检测多表格，如果为None则使用解析器的默认设置
            
        Returns:
            ExcelDocument对象
        """
        if detect_multiple_tables is None:
            detect_multiple_tables = self.detect_multiple_tables
        
        doc = ExcelDocument(file_path, detect_multiple_tables=detect_multiple_tables,
                          semantic_resolver=self.semantic_resolver)
        self.documents[doc.file_name] = doc
        return doc
    
    def load_documents(self, file_paths: List[str], detect_multiple_tables: Optional[bool] = None) -> List[ExcelDocument]:
        """
        批量加载多个Excel文档
        
        Args:
            file_paths: Excel文件路径列表
            detect_multiple_tables: 是否检测多表格，如果为None则使用解析器的默认设置
            
        Returns:
            ExcelDocument对象列表
        """
        return [self.load_document(path, detect_multiple_tables) for path in file_paths]
    
    def get_document(self, file_name: str, use_semantic: bool = True) -> Optional[ExcelDocument]:
        """
        获取已加载的文档
        
        Args:
            file_name: 文件名
            use_semantic: 是否使用语义匹配（如果精确匹配失败）
            
        Returns:
            ExcelDocument对象，如果不存在则返回None
        """
        # Try exact match first
        if file_name in self.documents:
            return self.documents.get(file_name)
        
        # Try semantic matching
        if use_semantic and self.semantic_resolver:
            matched_doc = self.semantic_resolver.find_document_by_semantic(
                list(self.documents.keys()),
                file_name
            )
            if matched_doc:
                logger.info(f"Semantic document match: '{file_name}' -> '{matched_doc}'")
                return self.documents.get(matched_doc)
        
        return None
    
    def get_all_documents(self) -> Dict[str, ExcelDocument]:
        """获取所有已加载的文档"""
        return self.documents
    
    def get_documents_summary(self) -> Dict[str, Any]:
        """
        获取所有文档的摘要信息
        
        Returns:
            包含所有文档统计信息的字典
        """
        return {
            "total_documents": len(self.documents),
            "documents": {
                name: doc.get_document_summary() 
                for name, doc in self.documents.items()
            }
        }
    
    def search_across_documents(self, value: Any) -> Dict[str, List[Dict[str, Any]]]:
        """
        在所有文档中搜索值
        
        Args:
            value: 要搜索的值
            
        Returns:
            按文档名分组的搜索结果
        """
        results = {}
        for name, doc in self.documents.items():
            doc_results = doc.search_value(value)
            if doc_results:
                results[name] = doc_results
        return results
    
    def join_sheets(self, doc1_name: str, sheet1_name: str, 
                   doc2_name: str, sheet2_name: str,
                   left_on: Optional[str] = None, right_on: Optional[str] = None, 
                   how: str = 'inner', auto_discover: bool = True) -> Optional[pd.DataFrame]:
        """
        跨文档关联工作表（支持自动发现关联键）
        
        Args:
            doc1_name: 第一个文档名
            sheet1_name: 第一个工作表名
            doc2_name: 第二个文档名
            sheet2_name: 第二个工作表名
            left_on: 第一个表的关联键（可选，如果为None则自动发现）
            right_on: 第二个表的关联键（可选，如果为None则自动发现）
            how: 关联方式 ('inner', 'left', 'right', 'outer')
            auto_discover: 是否自动发现关联键（当left_on/right_on未指定时）
            
        Returns:
            关联后的DataFrame
        """
        doc1 = self.get_document(doc1_name)
        doc2 = self.get_document(doc2_name)
        
        if not doc1 or not doc2:
            logger.error("文档未找到")
            return None
        
        df1 = doc1.get_sheet(sheet1_name)
        df2 = doc2.get_sheet(sheet2_name)
        
        if df1 is None or df2 is None:
            logger.error("工作表未找到")
            return None
        
        # Auto-discover join keys if not provided
        if (left_on is None or right_on is None) and auto_discover:
            logger.info("自动发现关联键...")
            discovered_keys = self.semantic_resolver.auto_discover_join_keys(df1, df2)
            if discovered_keys:
                left_on, right_on = discovered_keys
                logger.info(f"自动发现的关联键: left_on='{left_on}', right_on='{right_on}'")
            else:
                logger.error("无法自动发现关联键，请手动指定 left_on 和 right_on")
                return None
        
        if left_on is None or right_on is None:
            logger.error("未指定关联键且自动发现失败")
            return None
        
        return self.joiner.join_tables(df1, df2, left_on, right_on, how)
