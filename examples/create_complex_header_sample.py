"""
创建包含复杂表头的示例Excel文件
Create sample Excel files with complex headers (multi-row and hierarchical).
"""

from pathlib import Path
import openpyxl
from openpyxl import Workbook

examples_dir = Path(__file__).parent


def create_complex_header_sample():
    """创建包含复杂表头（多行+合并单元格）的示例"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "区域业绩统计"
    
    # 添加表格描述
    ws['A1'] = "表1：2024年各区域销售业绩统计"
    
    # 创建多行表头（3行）
    # 第一行：大区
    ws['A3'] = "区域"
    ws['B3'] = "华东"
    ws['D3'] = "华北"
    ws['F3'] = "华南"
    
    # 合并第一行的大区单元格
    ws.merge_cells('B3:C3')
    ws.merge_cells('D3:E3')
    ws.merge_cells('F3:G3')
    
    # 第二行：城市
    ws['A4'] = "城市"
    ws['B4'] = "南京"
    ws['C4'] = "苏州"
    ws['D4'] = "北京"
    ws['E4'] = "天津"
    ws['F4'] = "广州"
    ws['G4'] = "深圳"
    
    # 第三行：区县
    ws['A5'] = "具体区域"
    ws['B5'] = "浦口"
    ws['C5'] = "昆山"
    ws['D5'] = "朝阳"
    ws['E5'] = "滨海"
    ws['F5'] = "天河"
    ws['G5'] = "南山"
    
    # 合并第一列的三行
    ws.merge_cells('A3:A5')
    
    # 添加数据行
    data = [
        ["销售额(万元)", 1200, 980, 1500, 850, 1100, 1300],
        ["利润率(%)", 18.5, 16.2, 19.8, 15.5, 17.3, 20.1],
        ["客户数", 245, 198, 312, 156, 221, 289],
        ["增长率(%)", 12.3, 8.7, 15.2, 6.5, 10.8, 18.6],
    ]
    
    for i, row in enumerate(data, start=6):
        for j, value in enumerate(row):
            ws.cell(row=i, column=j+1, value=value)
    
    # 保存文件
    output_path = examples_dir / 'complex_header_sample.xlsx'
    wb.save(output_path)
    print(f"创建复杂表头示例文件: {output_path}")
    print("  - 3行表头（区域-城市-区县）")
    print("  - 包含合并单元格")
    print("  - 应解析为: 区域, 华东-南京-浦口, 华东-苏州-昆山, 华北-北京-朝阳, ...")


def create_multi_table_with_small_data():
    """创建包含多个表格和小型临时数据的示例"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "多表格混合"
    
    # 表1：主要销售数据（大表格）
    ws['A1'] = "表1：2024年主要销售数据"
    
    # 表头
    headers1 = ['产品', '销售额', '成本', '利润', '利润率']
    for col_idx, header in enumerate(headers1, 1):
        ws.cell(row=3, column=col_idx, value=header)
    
    # 数据（5行）
    data1 = [
        ['笔记本电脑', 500000, 350000, 150000, '30%'],
        ['台式电脑', 320000, 224000, 96000, '30%'],
        ['显示器', 150000, 105000, 45000, '30%'],
        ['键盘', 80000, 48000, 32000, '40%'],
        ['鼠标', 60000, 36000, 24000, '40%'],
    ]
    
    for row_idx, row_data in enumerate(data1, start=4):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 小型临时数据1（应该被过滤掉 - 只有1行数据）
    ws['G1'] = "临时备注"
    ws['G2'] = "已审核"
    
    # 小型临时数据2（应该被过滤掉 - 只有2行数据，少于MIN_TABLE_ROWS=3）
    ws['G5'] = "状态"
    ws['H5'] = "负责人"
    ws['G6'] = "进行中"
    ws['H6'] = "张三"
    ws['G7'] = "已完成"
    ws['H7'] = "李四"
    
    # 表2：费用统计（中等表格，应该被检测）
    ws['A12'] = "表2：2024年费用统计"
    
    headers2 = ['费用类型', '金额', '占比', '同比']
    for col_idx, header in enumerate(headers2, 1):
        ws.cell(row=14, column=col_idx, value=header)
    
    # 数据（4行，大于等于MIN_TABLE_ROWS）
    data2 = [
        ['人工成本', 250000, '50%', '+8%'],
        ['租金', 100000, '20%', '+5%'],
        ['营销费用', 80000, '16%', '+12%'],
        ['其他', 70000, '14%', '+3%'],
    ]
    
    for row_idx, row_data in enumerate(data2, start=15):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    output_path = examples_dir / 'multi_table_with_small_data.xlsx'
    wb.save(output_path)
    print(f"\n创建多表格混合示例文件: {output_path}")
    print("  - 2个有效表格（满足最小尺寸要求）")
    print("  - 2个小型临时数据（应被过滤）")


if __name__ == '__main__':
    create_complex_header_sample()
    create_multi_table_with_small_data()
    print("\n复杂表头示例文件创建完成！")
