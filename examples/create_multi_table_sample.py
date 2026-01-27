"""
创建包含多表格的示例Excel文件
Create sample Excel files with multiple tables in one sheet.
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

examples_dir = Path(__file__).parent


def create_multi_table_sheet():
    """创建一个sheet包含多个表格的示例"""
    
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "多表格汇总"
    
    # 添加第一个表格的描述（在A1）
    ws['A1'] = "表1：2024年Q1销售数据"
    
    # 第一个表格：从A3开始
    table1_data = {
        '产品': ['笔记本电脑', '台式电脑', '显示器'],
        '销售量': [120, 80, 150],
        '单价': [5000, 4000, 1500],
        '销售额': [600000, 320000, 225000]
    }
    table1_df = pd.DataFrame(table1_data)
    
    # 写入第一个表格（从第3行开始）
    for r_idx, row in enumerate(dataframe_to_rows(table1_df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 添加第二个表格的描述（在A9，有一行空白）
    ws['A9'] = "表2：2024年Q1费用明细"
    
    # 第二个表格：从A11开始
    table2_data = {
        '费用类型': ['人工成本', '租金', '水电费', '办公用品'],
        '金额': [150000, 50000, 8000, 12000],
        '占比': ['62.5%', '20.8%', '3.3%', '5.0%']
    }
    table2_df = pd.DataFrame(table2_data)
    
    # 写入第二个表格（从第11行开始）
    for r_idx, row in enumerate(dataframe_to_rows(table2_df, index=False, header=True), 11):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 添加第三个表格的描述（在F1，右侧偏移）
    ws['F1'] = "表3：客户分布"
    
    # 第三个表格：从F3开始（有列偏移）
    table3_data = {
        '区域': ['华北', '华东', '华南', '西部'],
        '客户数': [45, 78, 62, 35]
    }
    table3_df = pd.DataFrame(table3_data)
    
    # 写入第三个表格（从第3行，F列开始）
    for r_idx, row in enumerate(dataframe_to_rows(table3_df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 6):  # 从F列（第6列）开始
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 保存文件
    output_path = examples_dir / 'multi_table_sheet.xlsx'
    wb.save(output_path)
    print(f"创建多表格示例文件: {output_path}")


def create_offset_table_sheet():
    """创建一个有偏移量的单表格示例"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "偏移表格"
    
    # 添加标题和描述（在C2）
    ws['C2'] = "财务数据总览"
    ws['C3'] = "以下为2024年度财务指标"
    
    # 表格从C5开始（有左侧和上方偏移）
    table_data = {
        '指标': ['营业收入', '营业成本', '毛利率', '净利率'],
        '2023年': [15000000, 9000000, '40%', '18%'],
        '2024年': [18000000, 10500000, '42%', '20%'],
        '同比增长': ['+20%', '+17%', '+2pp', '+2pp']
    }
    table_df = pd.DataFrame(table_data)
    
    # 写入表格（从第5行，C列开始）
    for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=True), 5):
        for c_idx, value in enumerate(row, 3):  # 从C列（第3列）开始
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 保存文件
    output_path = examples_dir / 'offset_table_sheet.xlsx'
    wb.save(output_path)
    print(f"创建偏移表格示例文件: {output_path}")


if __name__ == '__main__':
    create_multi_table_sheet()
    create_offset_table_sheet()
    print("\n多表格示例文件创建完成！")
